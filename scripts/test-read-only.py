"""
Script para probar read_only=True vs sin read_only
"""
import openpyxl
import time
from datetime import datetime

xlsx_path = r'C:\STC\rptAcompDiario.xlsx'
sheet_name = 'report5'

print("=" * 60)
print("TEST 1: read_only=True, data_only=True")
print("=" * 60)
start = time.time()
wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
ws = wb[sheet_name]
print(f"Max columnas: {ws.max_column}")
print(f"Max filas: {ws.max_row}")

# Buscar registros RETALHO del 02/12/2025
count = 0
valores = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row and len(row) > 18 and row[1] and row[18] == 'RETALHO':
        if isinstance(row[1], datetime) and row[1].date() == datetime(2025, 12, 2).date():
            count += 1
            if len(row) > 17:
                valores.append(row[17])

wb.close()
elapsed = time.time() - start

print(f"Registros encontrados: {count}")
print(f"Valores METRAGEM: {valores[:5]}")
print(f"Tiempo: {elapsed:.2f}s")

print("\n" + "=" * 60)
print("TEST 2: Solo data_only=True (sin read_only)")
print("=" * 60)
start = time.time()
wb = openpyxl.load_workbook(xlsx_path, data_only=True)
ws = wb[sheet_name]
print(f"Max columnas: {ws.max_column}")
print(f"Max filas: {ws.max_row}")

# Buscar registros RETALHO del 02/12/2025
count = 0
valores = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row and row[1] and row[18] == 'RETALHO':
        if isinstance(row[1], datetime) and row[1].date() == datetime(2025, 12, 2).date():
            count += 1
            valores.append(row[17])

wb.close()
elapsed = time.time() - start

print(f"Registros encontrados: {count}")
print(f"Valores METRAGEM: {valores}")
print(f"Total suma: {sum(valores)}")
print(f"Tiempo: {elapsed:.2f}s")
